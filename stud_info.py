import pymysql
import openpyxl
import xlsxwriter


class Student:

    def __init__(self, sid, snm, sag, smail, smarks):
        self.stud_id = sid
        self.stud_name = snm
        self.stud_age = sag
        self.stud_email = smail
        self.stud_marks = smarks

    def __str__(self):
        return f'''\n {self.__dict__}'''

    def __repr__(self):
        return str(self)


def take_input_from_user():
    sid = int(input('Enter Stud Id : '))
    snm = input('Enter Stud Name : ')
    sag = int(input('Enter Stud Age : '))
    smail = input("Enter the Stud mail: ")
    smarks = float(input('Enter Stud marks : '))

    return Student(sid=sid, snm=snm, sag=sag, smail=smail, smarks=smarks)


CREATE_TABLE = '''
        CREATE TABLE STUD_INFO (
        Stud_ID INT,
        Stud_NAME VARCHAR(30),
        Stud_AGE INT,
        Stud_EMAIL VARCHAR(30),
        Stud_MARKS FLOAT,
        PRIMARY KEY(Stud_ID))
'''
studlist = []

while True:
    choice = input('Do you want to add more students in database : n/no')
    if choice.lower() in ['n', 'no']:
        break
    stud = take_input_from_user()
    studlist.append(stud)


# database lib --> excel lib...
# import pymysql
connection = pymysql.connect(host='localhost', user='root', password='root', port=3306, db='Student')

while True:
    print('''
            1. Export data into Database
            2. Export into Excel
            3. Import from DB To Excel
            4. Import Excel to DB
            5. CREATE TABLE
            6. Drop Table
            7. Exit
    ''')

    ch = int(input('Enter your Choice : '))
    if ch == 1:
        for stud in studlist:
            try:
                INSERT_QUERY = f''' 
                        INSERT INTO STUD_INFO VALUES({stud.stud_id},'{stud.stud_name}',{stud.stud_age},'{stud.stud_email}','{stud.stud_marks}')
                        '''
                channel = connection.cursor()
                channel.execute(INSERT_QUERY)
            except BaseException as e:
                print(e.args, stud.stud_id)
            else:
                connection.commit()
                print(f'{stud.stud_id} Saved Successfully....!', )
            finally:
                channel.close()
    elif ch == 2:

        workbook = openpyxl.Workbook()  # it will create one workbook for you..
        sheet1 = workbook.create_sheet('STUD-DATA-1')

        # headers -->
        sheet1['A1'] = 'STUD_ID'
        sheet1['B1'] = 'STUD_NAME'
        sheet1['C1'] = 'STUD_AGE'
        sheet1['D1'] = 'STUD_EMAIL'
        sheet1['E1'] = 'STUD_Marks'
        print('Headers Created...')

        row_num = 2
        for stud in studlist:
            sheet1['A' + str(row_num)] = stud.stud_id
            sheet1['B' + str(row_num)] = stud.stud_name
            sheet1['C' + str(row_num)] = stud.stud_age
            sheet1['D' + str(row_num)] = stud.stud_email
            sheet1['E' + str(row_num)] = stud.stud_marks

            row_num = row_num + 1

        workbook.save('student.xlsx')
        workbook.close()

    elif ch == 3:

        def fetch_table_data(stud_info):
            connection1 = pymysql.connect(host='localhost', user='root', password='root', port=3306, db='Student')
            channel1 = connection1.cursor()
            channel1.execute('select * from stud_info')
            header = [row[0] for row in channel1.description]

            rows = channel1.fetchall()
            connection1.close()

            return header, rows

        def export(stud_info):
            workbook = xlsxwriter.Workbook(stud_info + '.xlsx')
            worksheet = workbook.add_worksheet('STUD-DATA-1')

            header, rows = fetch_table_data(stud_info)
            row_index = 0
            column_index = 0

            for column_name in header:
                worksheet.write(row_index, column_index, column_name)
                column_index += 1

            row_index += 1
            for row in rows:
                column_index = 0
            for column in row:
                worksheet.write(row_index, column_index, column)
                column_index += 1
            row_index += 1

            print(str(row_index) + ' rows written successfully to ' + workbook.filename)

    # Closing workbook
            workbook.close()


# Tables to be exported
        export('stud_info')


    elif ch == 4:
        workbook = openpyxl.load_workbook('student.xlsx')
        sheet = workbook['STUD-DATA-1']
        num = sheet.max_row
        stud_list_from_excel = []
        for i in range(1,num+1):
            if i == 1:
                continue
            stud_id = int(sheet['A'+str(i)].value)
            stud_name = sheet['B'+str(i)].value
            stud_age = int(sheet['C'+str(i)].value)
            stud_mail = sheet['D'+str(i)].value
            stud_marks = float(sheet['E'+str(i)].value)
            stud=Student(sid=stud_id, snm=stud_name, sag=stud_age, smail=stud_mail, smarks=stud_marks)
            stud_list_from_excel.append(stud)
        print(stud_list_from_excel)  # read from excel
        for stud in stud_list_from_excel: #write into excel
            try:
                INSERT_QUERY = f''' 
                        INSERT INTO STUD_INFO VALUES({stud.stud_id},'{stud.stud_name}',{stud.stud_age},'{stud.stud_email}','{stud.stud_marks}')
                        '''
                channel = connection.cursor()
                channel.execute(INSERT_QUERY)
            except BaseException as e:
                print(e.args, stud.stud_id)
            else:
                connection.commit()
                print(f'{stud.stud_id} Saved Successfully....!', )
            finally:
                channel.close()

    elif ch == 5:
        try:
            channel = connection.cursor()
            channel.execute(CREATE_TABLE)
        except BaseException as e:
            print(e.args)
        else:
            connection.commit()
            print('Table Created Successfully...!')
        finally:
            channel.close()

    elif ch == 6:
        try:
            channel = connection.cursor()
            channel.execute('Drop table emp_info')
        except BaseException as e:
            print(e.args)
        else:
            connection.commit()
            print('Table Created Successfully...!')
        finally:
            channel.close()
    else:
        choice = input('Do you want to continue : n/no')
        if choice.lower() in ['n', 'no']:
            break
print('Invalid Option Selected...')
